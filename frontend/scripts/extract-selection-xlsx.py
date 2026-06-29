"""선발 명단 xlsx의 각 시트에서 (이름, 직위, 소속) 추출 → JSON."""
import json
import openpyxl

path = r"C:\Users\USER\Downloads\1. 2026년 AI챔피언 및 공공 AI역량 트랙 교육 (6월2차) 선발 명단_수정 (1).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

# 시트 이름 → cohort 이름 매핑 (예측)
SHEET_TO_COHORT = {
    'AI챔피언 그린 선발결과 1회차(6.16.~6.29.)': 'AI 챔피언 그린 1회차',
    'AI챔피언 블루 선발결과 2회차(6.23.~7.6.)': 'AI 챔피언 블루 2회차',
    '공공 AI역량 트랙 - ①AI리터러시와~(6.15)': 'AI 리터러시와 업무활용',
    '공공 AI역량 트랙 - ②생성형 AI 활용~(6.16)': '생성형 AI 활용 노코드 데이터분석',
    '공공 AI역량 트랙 - ③데이터 리터러시(6.17)': '데이터 리터러시',
    '공공 AI역량 트랙 - ④AI 행정~(6.29~6.30)': 'AI 행정 융합 기획'
}

output = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    # 데이터 시작 row 찾기 — '번호' 헤더 다음
    data_start = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip() == '번호':
            data_start = i + 1
            break
    if data_start is None:
        continue
    students = []
    for r in rows[data_start:]:
        if r[0] is None or str(r[0]).strip() == '':
            continue
        name = (r[1] if len(r) > 1 else '') or ''
        title = (r[2] if len(r) > 2 else '') or ''
        org = (r[3] if len(r) > 3 else '') or ''
        if not str(name).strip():
            continue
        students.append({
            'name': str(name).strip(),
            'title': str(title).strip(),
            'org': str(org).strip()
        })
    output[sn] = students
    print(f"{sn}: {len(students)}명 추출")

with open('/tmp/selection-xlsx.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False)
print('\n저장: /tmp/selection-xlsx.json')
