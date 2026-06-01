"""
2025년 AI챔피언 인증자 명단 → applicants.prior_certs 적재.

대상 파일: 별첨. AI챔피언 인증자 명단 원본 데이터(개인정보 포함) _data2025!.xlsx (12 시트, 541명)
매칭 우선순위: name + birth_date (둘 다 있을 때) → email (대체)
신규 응시자는 INSERT, 기존 매칭 시 prior_certs 배열에 머지(cert_no 기준 dedupe).

선행 조건:
  마이그레이션 20260601000001_applicants_prior_certs.sql 적용 완료

실행:
  cd frontend
  PYTHONIOENCODING=utf-8 python -X utf8 scripts/import-2025-certifications.py
"""
import io
import os
import re
import sys
from datetime import date, datetime, timedelta
from typing import Any
import urllib.request
import urllib.error
import json

import msoffcrypto
from openpyxl import load_workbook

FILE_A = r'C:\Users\USER\Documents\카카오톡 받은 파일\별첨. AI챔피언 인증자 명단 원본 데이터(개인정보 포함) _data2025!.xlsx'
PASSWORD = 'data2025!'

CERT_NO_RE = re.compile(r'^2025-[A-Z]{2}\d{4}$')

# sheet name → (track, round, event)
SHEET_MAP: dict[str, tuple[str, int | None, str | None]] = {
    '그린 1회차 인증': ('green', 1, None),
    '그린 2회차 인증': ('green', 2, None),
    '그린 3회차 인증': ('green', 3, None),
    '블루 1회차 인증': ('blue', 1, None),
    '블루 2회차 인증': ('blue', 2, None),
    '블루 3회차 인증': ('blue', 3, None),
    '블루 4회차 인증 민간협업': ('blue', 4, 'private'),
    '전문인재 인증': ('expert', None, None),
    '보수교육 인증': ('continuing', None, None),
    '헤커톤 2회차(블루)': ('blue', 2, 'hackathon'),
    '미니프로젝트 2회차(블루)': ('blue', 2, 'miniproject'),
}


# ---------------------------------------------------------------- env / http
def load_env():
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
    with open(env_path, encoding='utf-8') as f:
        for line in f:
            m = re.match(r'^([A-Z_]+)=(.*)$', line.strip())
            if m:
                os.environ[m.group(1)] = m.group(2).strip('"')


def supabase_request(method: str, path: str, body: Any = None, params: dict | None = None):
    url = os.environ['NEXT_PUBLIC_SUPABASE_URL'] + '/rest/v1' + path
    if params:
        from urllib.parse import urlencode
        url += ('&' if '?' in url else '?') + urlencode(params)
    headers = {
        'apikey': os.environ['SUPABASE_SERVICE_ROLE_KEY'],
        'Authorization': 'Bearer ' + os.environ['SUPABASE_SERVICE_ROLE_KEY'],
        'Content-Type': 'application/json',
        'Prefer': 'return=representation',
    }
    data = json.dumps(body).encode('utf-8') if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode('utf-8')
            return json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        body_err = e.read().decode('utf-8', errors='replace')
        raise RuntimeError(f'{method} {url} → {e.code}: {body_err}') from e


# ---------------------------------------------------------------- xlsx parse
def s(v):
    if v is None:
        return ''
    return str(v).strip()


def norm_birth(v) -> str:
    if v is None or v == '':
        return ''
    if isinstance(v, (int, float)) and 10000 < float(v) < 80000:
        d = date(1899, 12, 30) + timedelta(days=int(v))
        return d.strftime('%Y%m%d')
    if isinstance(v, (date, datetime)):
        return v.strftime('%Y%m%d')
    digits = re.sub(r'\D', '', str(v).strip())
    return digits[:8] if len(digits) >= 8 else ''


def birth_to_iso(b8: str) -> str | None:
    if len(b8) != 8 or not b8.isdigit():
        return None
    return f'{b8[0:4]}-{b8[4:6]}-{b8[6:8]}'


def norm_email_list(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r'[,;\s]+', raw)
    return [p.strip().lower() for p in parts if '@' in p]


def find_col(header, *candidates):
    for idx, h in enumerate(header, start=1):
        ht = s(h)
        if not ht:
            continue
        for c in candidates:
            if c in ht:
                return idx
    return None


def open_workbook():
    buf = io.BytesIO()
    with open(FILE_A, 'rb') as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=PASSWORD)
        office.decrypt(buf)
    buf.seek(0)
    return load_workbook(buf, data_only=True, read_only=True)


def collect_cert_rows():
    """모든 시트에서 cert_no가 있는 행을 단일 리스트로 모음"""
    wb = open_workbook()
    out = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_MAP:
            continue
        ws = wb[sheet_name]
        # 헤더 행 찾기
        header_row = None
        header = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if any('인증번호' in s(v) for v in row):
                header_row = r_idx
                header = list(row)
                break
        if header_row is None:
            continue
        col_cert = find_col(header, '인증번호')
        col_name = find_col(header, '이름')
        col_org = find_col(header, '기관')
        col_birth = find_col(header, '생년월일')
        # 시트마다 '메일'/'이메일'/'회사메일' 가능
        col_email = find_col(header, '메일')
        col_email_alt = None
        if col_email:
            # '회사메일'이 잡혔으면 '개인메일'도 따로
            ht = s(header[col_email - 1])
            if '회사' in ht or '개인' in ht:
                # 둘 다 찾기
                col_email = find_col(header, '회사메일') or col_email
                col_email_alt = find_col(header, '개인메일')
        col_category = find_col(header, '분류')
        col_kind = find_col(header, '구분')
        if not col_cert or not col_name:
            continue
        track, rnd, event = SHEET_MAP[sheet_name]
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cert = s(row[col_cert - 1]) if col_cert <= len(row) else ''
            if not cert or not CERT_NO_RE.match(cert):
                continue
            emails = []
            if col_email and col_email <= len(row):
                emails.extend(norm_email_list(s(row[col_email - 1])))
            if col_email_alt and col_email_alt <= len(row):
                emails.extend(norm_email_list(s(row[col_email_alt - 1])))
            # dedupe 유지 순서
            seen = set()
            emails = [e for e in emails if not (e in seen or seen.add(e))]
            out.append({
                'sheet': sheet_name,
                'cert_no': cert,
                'name': s(row[col_name - 1]),
                'org': s(row[col_org - 1]) if col_org else '',
                'birth': norm_birth(row[col_birth - 1]) if col_birth else '',
                'emails': emails,
                'category': s(row[col_category - 1]) if col_category else '',
                'kind': s(row[col_kind - 1]) if col_kind else '',
                'track': track,
                'round': rnd,
                'event': event,
            })
    return out


# ---------------------------------------------------------------- import
def main():
    load_env()
    print('Loading xlsx...', file=sys.stderr)
    rows = collect_cert_rows()
    print(f'Found {len(rows)} cert rows', file=sys.stderr)

    # (name, birth) → list[row]
    person_keys: dict[tuple[str, str], list[dict]] = {}
    for r in rows:
        key = (r['name'], r['birth'])
        person_keys.setdefault(key, []).append(r)
    print(f'Unique people: {len(person_keys)}', file=sys.stderr)

    # 기존 applicants 페이지네이션 조회 (PostgREST 기본 limit=1000)
    print('Fetching existing applicants...', file=sys.stderr)
    existing: list[dict] = []
    offset = 0
    page = 1000
    while True:
        batch = supabase_request('GET', '/applicants', params={
            'select': 'id,name,birth_date,email,prior_certs',
            'limit': page,
            'offset': offset,
            'order': 'id.asc',
        })
        if not batch:
            break
        existing.extend(batch)
        if len(batch) < page:
            break
        offset += page
    print(f'  → {len(existing)} existing applicants', file=sys.stderr)
    by_name_birth: dict[tuple[str, str], dict] = {}
    by_email: dict[str, dict] = {}
    for a in existing or []:
        nb = a.get('birth_date')
        if a.get('name') and nb:
            by_name_birth[(a['name'], nb)] = a
        if a.get('email'):
            by_email[a['email'].strip().lower()] = a

    matched_existing = 0
    new_inserts = 0
    cert_added = 0
    cert_skipped_dup = 0

    for (name, birth8), cert_rows in person_keys.items():
        birth_iso = birth_to_iso(birth8)
        emails = []
        for r in cert_rows:
            for e in r['emails']:
                if e not in emails:
                    emails.append(e)
        primary_email = emails[0] if emails else None

        # 매칭
        match = None
        if birth_iso and (name, birth_iso) in by_name_birth:
            match = by_name_birth[(name, birth_iso)]
        else:
            for e in emails:
                if e in by_email:
                    match = by_email[e]
                    break

        # cert objects 빌드
        track_label = {'green': '그린', 'blue': '블루', 'expert': '전문인재', 'continuing': '보수교육'}
        new_certs = []
        for r in cert_rows:
            new_certs.append({
                'year': 2025,
                'cert_no': r['cert_no'],
                'track': r['track'],
                'round': r['round'],
                'kind': r['kind'] or None,
                'event': r['event'],
                'organization': r['org'] or None,
                'cert_name': f'2025년 AI챔피언 {track_label[r["track"]]}',
            })

        if match:
            matched_existing += 1
            existing_certs = match.get('prior_certs') or []
            existing_nos = {c.get('cert_no') for c in existing_certs}
            additions = []
            for c in new_certs:
                if c['cert_no'] in existing_nos:
                    cert_skipped_dup += 1
                    continue
                additions.append(c)
                cert_added += 1
            if additions:
                merged = list(existing_certs) + additions
                supabase_request('PATCH', f'/applicants?id=eq.{match["id"]}', body={'prior_certs': merged})
        else:
            new_inserts += 1
            cert_added += len(new_certs)
            payload = {
                'name': name,
                'birth_date': birth_iso,
                'email': primary_email,
                'prior_certs': new_certs,
            }
            supabase_request('POST', '/applicants', body=payload)

    print()
    print(f'=== 결과 ===')
    print(f'  XLS rows:        {len(rows)}')
    print(f'  Unique people:   {len(person_keys)}')
    print(f'  기존 매칭:        {matched_existing}명')
    print(f'  신규 등록:        {new_inserts}명')
    print(f'  추가된 인증:      {cert_added}건')
    print(f'  중복 skip:        {cert_skipped_dup}건')


if __name__ == '__main__':
    main()
