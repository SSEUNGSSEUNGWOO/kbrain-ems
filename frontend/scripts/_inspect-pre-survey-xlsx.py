"""Inspect the three updated pre-survey xlsx files and print all sheet contents."""
from pathlib import Path

import openpyxl

FILES = [
    r"C:\Users\USER\Downloads\자동화용 시트_2026년 AI 챔피언 그린(초급) 종합과정 2회차.xlsx",
    r"C:\Users\USER\Downloads\자동화용 시트_2026년 AI 챔피언 블루(중급) 종합과정 3회차.xlsx",
    r"C:\Users\USER\Downloads\자동화용 시트_2026년 ⑥ 노코드 AI 서비스 구현.xlsx",
]

OUT = Path(r"C:\Dev\kbrain\kbrain-ems\frontend\scripts\_pre-survey-dump.txt")

lines: list[str] = []
for path in FILES:
    lines.append("=" * 100)
    lines.append(f"FILE: {Path(path).name}")
    lines.append("=" * 100)
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            non_empty = [(i, c) for i, c in enumerate(row) if c is not None and str(c).strip()]
            if not non_empty:
                continue
            lines.append(f"  [{r_idx:>3}] " + " | ".join(f"<{i}>{c}" for i, c in non_empty))
    lines.append("")

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {OUT} ({len(lines)} lines)")
