"""선발 명단 xlsx 파일 구조 + 데이터 추출."""
import sys, os, json
import openpyxl

path = r"C:\Users\USER\Downloads\1. 2026년 AI챔피언 및 공공 AI역량 트랙 교육 (6월2차) 선발 명단_수정 (1).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print("=== 시트 목록 ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  '{sn}': {ws.max_row} rows × {ws.max_column} cols")

# 각 시트의 첫 2행 (헤더 + 첫 데이터) 표시
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== '{sn}' 헤더 ===")
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    for i, r in enumerate(rows):
        print(f"  row {i}: {' | '.join(str(c) if c is not None else '' for c in r[:15])}{' ...' if len(r) > 15 else ''}")
